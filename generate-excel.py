import json
import base64
from io import BytesIO

def handler(event, context):
    try:
        # 解析請求
        body = json.loads(event.get('body', '{}'))
        name = body.get('name', '業主')
        today = body.get('today', '2026/1/1')
        sections = body.get('sections', [])

        # 生成 Excel
        b64 = gen_excel(name, today, sections)

        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Headers': 'Content-Type',
            },
            'body': json.dumps({'b64': b64})
        }
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': str(e)})
        }

def gen_excel(name, today, sections_data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    BLUEH='4472C4'; WHITE='FFFFFF'; YELLOW='FFFF00'; RED='C00000'
    NUMS=['一','二','三','四','五','六','七','八','九','十','十一','十二','十三','十四']

    def sd(s='thin',c='000000'): return Side(style=s, color=c)
    def ball(s='thin',c='000000'):
        x=sd(s,c)
        return Border(left=x, right=x, top=x, bottom=x)
    def ft(bold=False, sz=11, color='000000'):
        return Font(bold=bold, sz=sz, color=color, name='新細明體')
    def al(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def fill(c):
        return PatternFill('solid', fgColor=c)

    wb = Workbook()
    ws = wb.active
    ws.title = '澤居報價單'

    for col, w in zip('ABCDEFG', [8, 32, 8, 8, 12, 12, 18]):
        ws.column_dimensions[col].width = w

    for r in range(1, 27):
        for c in range(1, 8):
            ws.cell(r, c).border = ball()

    # R1 公司名
    ws.row_dimensions[1].height = 38
    ws.merge_cells('A1:G1')
    ws['A1'].value = '澤\u3000居\u3000室\u3000內\u3000裝\u3000修'
    ws['A1'].font = ft(True, 16)
    ws['A1'].alignment = al('center')

    # R2 業主
    ws.row_dimensions[2].height = 20
    ws.merge_cells('A2:D2')
    ws['A2'].value = '業主：' + name
    ws['A2'].font = ft(False, 11)
    ws['A2'].alignment = al('left')
    ws.merge_cells('E2:F2')
    ws['E2'].value = '製表：'
    ws['E2'].font = ft(False, 11)
    ws['G2'].value = '陳鴻彬'
    ws['G2'].font = ft(False, 11)
    ws['G2'].alignment = al('center')

    # R3 地址/日期
    ws.row_dimensions[3].height = 20
    ws.merge_cells('A3:D3')
    ws['A3'].value = '地址：'
    ws['A3'].font = ft(False, 11)
    ws['A3'].alignment = al('left')
    ws.merge_cells('E3:F3')
    ws['E3'].value = '日期：' + today
    ws['E3'].font = ft(False, 11)
    ws['G3'].value = '報價有效期限15日'
    ws['G3'].font = Font(sz=10, color=RED, name='新細明體')
    ws['G3'].alignment = al('center')

    # R4 報價總單
    ws.row_dimensions[4].height = 16
    ws.merge_cells('A4:G4')
    ws['A4'].value = '報價總單'
    ws['A4'].font = ft(False, 11)
    ws['A4'].alignment = al('center')

    # R5 表頭
    ws.row_dimensions[5].height = 24
    for ci, h in enumerate(['項次','工程種類別','單位','數量','單價','複價','備註'], 1):
        c = ws.cell(5, ci)
        c.value = h
        c.font = ft(True, 12, WHITE)
        c.fill = fill(BLUEH)
        c.alignment = al('center')
        c.border = ball('medium')

    # 計算合計
    grand = 0
    sec_totals = []
    for sec in sections_data:
        t = sum(it.get('price', 0) * it.get('qty', 1) for it in sec.get('items', []))
        sec_totals.append(round(t))
        grand += round(t)
    tax = round(grand * 0.05)
    total_amt = grand + tax

    # R6-19 資料行
    for i in range(14):
        r = 6 + i
        ws.row_dimensions[r].height = 22
        sec = sections_data[i] if i < len(sections_data) else None
        st = sec_totals[i] if i < len(sec_totals) else 0
        ws.cell(r,1).value = NUMS[i]; ws.cell(r,1).font = ft(False,12); ws.cell(r,1).alignment = al('center')
        ws.cell(r,2).value = sec['name'] if sec else ''; ws.cell(r,2).font = ft(False,11); ws.cell(r,2).alignment = al('center')
        ws.cell(r,3).value = '式'; ws.cell(r,3).alignment = al('center')
        ws.cell(r,4).value = 1; ws.cell(r,4).alignment = al('center')
        ws.cell(r,6).value = st; ws.cell(r,6).font = ft(True, 12, RED); ws.cell(r,6).alignment = al('right')

    # R20-22 合計區
    ws.row_dimensions[20].height = 20
    ws.merge_cells('A20:E20')
    ws['F20'].value = grand; ws['F20'].font = ft(True, 12, RED); ws['F20'].alignment = al('right')

    ws.row_dimensions[21].height = 20
    ws.merge_cells('A21:E21')
    ws['A21'].value = '稅金5%'; ws['A21'].font = ft(False, 11); ws['A21'].alignment = al('center')
    ws['F21'].value = tax; ws['F21'].font = ft(True, 12, RED); ws['F21'].alignment = al('right')

    ws.row_dimensions[22].height = 24
    ws.merge_cells('A22:E22')
    ws['A22'].value = '合計'; ws['A22'].font = ft(True, 13); ws['A22'].alignment = al('center')
    ws['A22'].border = ball('medium')
    ws['F22'].value = total_amt; ws['F22'].font = ft(True, 13, RED); ws['F22'].alignment = al('right')
    ws['F22'].border = ball('medium')
    ws['G22'].border = ball('medium')

    # R23 備注
    ws.row_dimensions[23].height = 68
    ws.merge_cells('A23:G23')
    ws['A23'].value = ('備註：一. 付款方式：第一期訂金30%，第二期施工進場3天內30%，第三期完成7成30%，第四期尾款驗收後10%\n'
                       '二. 每期請款請於提出後3日內付清，否則保留停工之權利\n'
                       '三. 此報價單確認無誤後請簽名回傳，即轉為正式合約')
    ws['A23'].font = ft(False, 10)
    ws['A23'].alignment = al('left', 'top', True)

    ws.row_dimensions[24].height = 10
    ws.merge_cells('A24:G24')

    ws.row_dimensions[25].height = 18
    ws.merge_cells('A25:G25')
    ws['A25'].value = '匯款資訊：銀行代碼：050\u3000台灣企銀-八德分行\u3000戶名：澤居室內裝修\u3000帳號：7505400208531'
    ws['A25'].font = ft(True, 10)
    ws['A25'].alignment = al('left')

    ws.row_dimensions[26].height = 52
    ws.merge_cells('A26:C26')
    ws['A26'].value = '公司蓋章處'; ws['A26'].font = ft(True, 12); ws['A26'].alignment = al('center')
    ws.merge_cells('D26:G26')
    ws['D26'].value = '客戶回簽處'; ws['D26'].font = ft(True, 12); ws['D26'].alignment = al('center')

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.4; ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5; ws.page_margins.bottom = 0.5
    ws.print_area = 'A1:G26'

    # 細項 Sheets
    used_names = ['澤居報價單']
    for si, sec in enumerate(sections_data):
        sname = sec['name'][:8]
        if sname in used_names:
            sname = sname[:6] + str(si)
        used_names.append(sname)
        ws2 = wb.create_sheet(sname)

        for col, w in zip('ABCDEFG', [8, 36, 8, 8, 12, 12, 18]):
            ws2.column_dimensions[col].width = w

        items = sec.get('items', [])
        total_rows = 5 + max(len(items), 15)

        for r in range(1, total_rows + 1):
            for c in range(1, 8):
                ws2.cell(r, c).border = ball()

        ws2.row_dimensions[1].height = 20
        ws2.merge_cells('A1:D1')
        ws2['A1'].value = '業主：' + name; ws2['A1'].font = ft(False,11); ws2['A1'].alignment = al('left')
        ws2.merge_cells('E1:F1')
        ws2['E1'].value = '製表：'; ws2['E1'].font = ft(False,11)
        ws2['G1'].value = '日期：' + today; ws2['G1'].font = ft(False,11)

        ws2.row_dimensions[2].height = 20
        ws2.merge_cells('A2:G2')
        ws2['A2'].value = '地址：'; ws2['A2'].font = ft(False,11); ws2['A2'].alignment = al('left')

        ws2.row_dimensions[3].height = 16
        ws2.merge_cells('A3:G3')
        ws2['A3'].value = '報價細項表'; ws2['A3'].font = ft(False,11); ws2['A3'].alignment = al('center')

        ws2.row_dimensions[4].height = 24
        for ci, h in enumerate(['項次','工程種類別','單位','數量','單價','複價','備註'], 1):
            c = ws2.cell(4, ci)
            c.value = h; c.font = ft(True,12,WHITE); c.fill = fill(BLUEH)
            c.alignment = al('center'); c.border = ball('medium')

        ws2.row_dimensions[5].height = 22
        ws2.merge_cells('A5:E5')
        ws2['A5'].value = sec['name']; ws2['A5'].font = ft(True,12); ws2['A5'].fill = fill(YELLOW); ws2['A5'].alignment = al('center')
        ws2['F5'].value = '合計'; ws2['F5'].font = ft(True,11,RED); ws2['F5'].fill = fill(YELLOW); ws2['F5'].alignment = al('right')
        ws2['G5'].value = sec_totals[si]; ws2['G5'].font = ft(True,12,RED); ws2['G5'].fill = fill(YELLOW); ws2['G5'].alignment = al('right')

        row = 6
        for ii, it in enumerate(items):
            ws2.row_dimensions[row].height = 22
            try:
                qty = float(str(it.get('qty', 1)).replace('式','').replace('坪','') or 1)
            except:
                qty = 1
            price = it.get('price', 0) or 0
            ws2.cell(row,1).value = ii+1; ws2.cell(row,1).alignment = al('center')
            ws2.cell(row,2).value = it.get('name',''); ws2.cell(row,2).font = ft(False,10); ws2.cell(row,2).alignment = al('left', wrap=True)
            ws2.cell(row,3).value = it.get('unit',''); ws2.cell(row,3).alignment = al('center')
            ws2.cell(row,4).value = qty; ws2.cell(row,4).alignment = al('center')
            ws2.cell(row,5).value = price if price else None; ws2.cell(row,5).alignment = al('right')
            ws2.cell(row,6).value = round(qty * price); ws2.cell(row,6).alignment = al('right')
            ws2.cell(row,7).value = it.get('note',''); ws2.cell(row,7).font = ft(False,10); ws2.cell(row,7).alignment = al('left', wrap=True)
            row += 1

        while row <= total_rows:
            ws2.row_dimensions[row].height = 22
            row += 1

        ws2.page_setup.orientation = 'portrait'
        ws2.page_setup.paperSize = 9
        ws2.page_setup.fitToPage = True
        ws2.page_setup.fitToWidth = 1
        ws2.page_setup.fitToHeight = 1
        ws2.page_margins.left = 0.4; ws2.page_margins.right = 0.4
        ws2.page_margins.top = 0.5; ws2.page_margins.bottom = 0.5

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()
